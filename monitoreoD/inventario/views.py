from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from .models import MovimientoInventario
from .forms import MovimientoInventarioForm
from django.http import JsonResponse, HttpResponse
from usuarios.decorators import requiere_permiso
from django.contrib.auth.decorators import login_required
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

@login_required
@requiere_permiso('ver', 'inventario')
def lista_movimientos(request):
    search = request.GET.get('search', '')
    order_by = request.GET.get('order_by', '-fecha')
    per_page = request.GET.get('per_page', '10')
    
    movimientos = MovimientoInventario.objects.all()
    
    if search:
        movimientos = movimientos.filter(
            Q(producto__nombre__icontains=search) | 
            Q(tipo__icontains=search) | 
            Q(lote__icontains=search)
        )
    
    if order_by in ['fecha', '-fecha', 'tipo', '-tipo', 'cantidad', '-cantidad', 'producto__nombre', '-producto__nombre']:
        movimientos = movimientos.order_by(order_by)
    
    paginator = Paginator(movimientos, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # AJAX response for live search
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        movimientos_data = []
        for movimiento in page_obj:
            movimientos_data.append({
                'id': movimiento.pk,
                'fecha': movimiento.fecha.strftime('%d/%m/%Y %H:%M'),
                'tipo': movimiento.get_tipo_display(),
                'producto': movimiento.producto.nombre,
                'cantidad': int(movimiento.cantidad) if movimiento.cantidad == int(movimiento.cantidad) else float(movimiento.cantidad),
                'lote': movimiento.lote or '',
                'tipo_class': 'text-bg-success' if movimiento.tipo == 'ENTRADA' else 'text-bg-danger'
            })
        return JsonResponse({
            'movimientos': movimientos_data,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'page_number': page_obj.number,
            'num_pages': page_obj.paginator.num_pages
        })
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'order_by': order_by,
        'per_page': per_page,
    }
    
    return render(request, 'inventario/lista.html', context)

@login_required
@requiere_permiso('crear', 'inventario')
def crear_movimiento(request):
    if request.method == 'POST':
        form = MovimientoInventarioForm(request.POST)
        if form.is_valid():
            movimiento = form.save()
            messages.success(request, f'Movimiento de inventario creado exitosamente.')
            return redirect('inventario:inventario_lista')
    else:
        form = MovimientoInventarioForm()
    return render(request, 'inventario/crear.html', {'form': form})

@login_required
@requiere_permiso('editar', 'inventario')
def editar_movimiento(request, pk):
    movimiento = get_object_or_404(MovimientoInventario, pk=pk)
    if request.method == 'POST':
        form = MovimientoInventarioForm(request.POST, instance=movimiento)
        if form.is_valid():
            form.save()
            messages.success(request, f'Movimiento de inventario actualizado exitosamente.')
            return redirect('inventario:inventario_lista')
    else:
        form = MovimientoInventarioForm(instance=movimiento)
    return render(request, 'inventario/editar.html', {'form': form, 'movimiento': movimiento})



@login_required
def detalle_movimiento(request, pk):
    movimiento = get_object_or_404(MovimientoInventario, pk=pk)
    return render(request, 'inventario/detalle.html', {'movimiento': movimiento})

@login_required
@requiere_permiso('exportar_inventario')
def exportar_inventario_excel(request):
    search = request.GET.get('search', '')
    order_by = request.GET.get('order_by', '-fecha')
    
    movimientos = MovimientoInventario.objects.all()
    
    if search:
        movimientos = movimientos.filter(
            Q(producto__nombre__icontains=search) | 
            Q(tipo__icontains=search) | 
            Q(lote__icontains=search)
        )
    
    if order_by in ['fecha', '-fecha', 'tipo', '-tipo', 'cantidad', '-cantidad', 'producto__nombre', '-producto__nombre']:
        movimientos = movimientos.order_by(order_by)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos Inventario"
    
    headers = ['Fecha', 'Tipo', 'Producto', 'Cantidad', 'Lote', 'Proveedor', 'Observaciones']
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
    
    for movimiento in movimientos:
        ws.append([
            movimiento.fecha.strftime('%d/%m/%Y %H:%M'),
            movimiento.get_tipo_display(),
            movimiento.producto.nombre,
            float(movimiento.cantidad),
            movimiento.lote or '',
            movimiento.proveedor.razon_social if movimiento.proveedor else '',
            movimiento.observaciones or ''
        ])
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'inventario_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response