from django.shortcuts import render
from .models import Usuario
from .test_permisos import test_productos, test_sin_decorador
from .forms import UsuarioCreationForm, UsuarioChangeForm, PerfilUsuarioForm, CambiarContrasenaForm
from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
from django import forms
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import update_session_auth_hash
from .decorators import requiere_permiso
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

@login_required
def dashboard(request):
    return render(request, 'dashboard.html')

class PasswordResetForm(forms.Form):
    email = forms.EmailField(
        widget=forms.EmailInput(attrs={'class': 'form-control', 'placeholder': 'Ingresa tu email'})
    )

def password_reset_request(request):
    if request.method == 'POST':
        form = PasswordResetForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            try:
                usuario = Usuario.objects.get(email=email)
                # Generar token de recuperación
                from django.contrib.auth.tokens import default_token_generator
                from django.utils.http import urlsafe_base64_encode
                from django.utils.encoding import force_bytes
                from django.urls import reverse
                
                token = default_token_generator.make_token(usuario)
                uid = urlsafe_base64_encode(force_bytes(usuario.pk))
                
                # Construir URL de recuperación
                reset_url = request.build_absolute_uri(
                    reverse('usuarios:password_reset_confirm', kwargs={'uidb64': uid, 'token': token})
                )
                
                # Enviar email con instrucciones
                subject = 'Recuperación de Contraseña - Lilis Dulcería'
                message = f'''Hola {usuario.first_name or usuario.username},

Has solicitado recuperar tu contraseña para el sistema de Lilis Dulcería.

Haz clic en el siguiente enlace para restablecer tu contraseña:
{reset_url}

Este enlace es válido por 24 horas.

Si no solicitaste este cambio, ignora este mensaje.

Saludos,
Equipo de Lilis Dulcería'''
                
                try:
                    send_mail(
                        subject,
                        message,
                        settings.DEFAULT_FROM_EMAIL,
                        [email],
                        fail_silently=False,
                    )
                    messages.success(request, 'Se han enviado las instrucciones de recuperación a tu email.')
                    return redirect('usuarios:password_reset_done')
                except Exception as e:
                    messages.error(request, f'Error al enviar el email: {str(e)}')
            except Usuario.DoesNotExist:
                # Por seguridad, no revelamos si el email existe o no
                messages.success(request, 'Si el email existe en nuestro sistema, recibirás las instrucciones de recuperación.')
                return redirect('usuarios:password_reset_done')
    else:
        form = PasswordResetForm()
    
    return render(request, 'usuarios/password_reset.html', {'form': form})

@login_required
@requiere_permiso('editar', 'usuarios')
def editar_usuario(request, pk):
    usuario = get_object_or_404(Usuario, pk=pk)
    if request.method == 'POST':
        form = UsuarioChangeForm(request.POST, instance=usuario)
        if form.is_valid():
            usuario = form.save()
            messages.success(request, f'Usuario "{usuario.username}" actualizado exitosamente.')
            return redirect('usuarios:usuarios_lista')
    else:
        form = UsuarioChangeForm(instance=usuario)
    return render(request, 'usuarios/editar.html', {'form': form, 'usuario': usuario, 'title': 'Editar Usuario'})

@login_required
@requiere_permiso('crear', 'usuarios')
def crear_usuario(request):
    if request.method == 'POST':
        form = UsuarioCreationForm(request.POST)
        if form.is_valid():
            usuario = form.save()
            messages.success(request, f'Usuario "{usuario.username}" creado exitosamente.')
            return redirect('usuarios:usuarios_lista')
    else:
        form = UsuarioCreationForm()
    return render(request, 'usuarios/crear.html', {'form': form, 'title': 'Crear Usuario'})

@login_required
@requiere_permiso('ver', 'usuarios')
def lista_usuarios(request):
    search = request.GET.get('search', '')
    order_by = request.GET.get('order_by', 'username')
    per_page = request.session.get('usuarios_per_page', request.GET.get('per_page', '10'))
    
    # Guardar paginación en sesión
    if request.GET.get('per_page'):
        request.session['usuarios_per_page'] = request.GET.get('per_page')
        per_page = request.GET.get('per_page')
    
    usuarios = Usuario.objects.all()
    
    if search:
        usuarios = usuarios.filter(
            Q(username__icontains=search) | 
            Q(email__icontains=search) | 
            Q(first_name__icontains=search) | 
            Q(last_name__icontains=search) |
            Q(rol__icontains=search)
        )
    
    if order_by in ['username', '-username', 'email', '-email', 'date_joined', '-date_joined', 'rol', '-rol']:
        usuarios = usuarios.order_by(order_by)
    
    paginator = Paginator(usuarios, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # AJAX response for live search
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        usuarios_data = []
        for usuario in page_obj:
            usuarios_data.append({
                'id': usuario.pk,
                'username': usuario.username,
                'nombres': usuario.first_name,
                'apellidos': usuario.last_name,
                'email': usuario.email,
                'rol': usuario.get_rol_display(),
                'rol_class': 'bg-danger' if usuario.rol == 'ADMINISTRADOR' else 'text-bg-danger' if 'GERENTE' in usuario.rol else 'text-bg-warning' if 'SUPERVISOR' in usuario.rol else 'text-bg-info' if usuario.rol in ['OPERADOR_INVENTARIO', 'VENDEDOR', 'CONTADOR'] else 'text-bg-secondary',
                'estado': usuario.estado,
                'estado_class': 'text-bg-success' if usuario.estado == 'ACTIVO' else 'text-bg-secondary'
            })
        return JsonResponse({
            'usuarios': usuarios_data,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'page_number': page_obj.number,
            'num_pages': page_obj.paginator.num_pages
        })
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'order_by': order_by,
        'per_page': per_page,
    }
    
    return render(request, 'usuarios/lista.html', context)

@login_required
@requiere_permiso('desactivar', 'usuarios')
def toggle_usuario_estado(request, pk):
    usuario = get_object_or_404(Usuario, pk=pk)
    if request.method == 'POST':
        username = usuario.username
        nuevo_estado = 'INACTIVO' if usuario.estado == 'ACTIVO' else 'ACTIVO'
        usuario.estado = nuevo_estado
        usuario.save()
        accion = 'desactivado' if nuevo_estado == 'INACTIVO' else 'activado'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'Usuario "{username}" {accion} exitosamente.'})
        messages.success(request, f'Usuario "{username}" {accion} exitosamente.')
        return redirect('usuarios:usuarios_lista')
    return render(request, 'usuarios/toggle_estado.html', {'usuario': usuario})

@login_required
def detalle_usuario(request, pk):
    usuario = get_object_or_404(Usuario, pk=pk)
    return render(request, 'usuarios/detalle.html', {'usuario': usuario})

@login_required
def perfil_usuario(request):
    if request.method == 'POST':
        form = PerfilUsuarioForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil actualizado exitosamente.')
            return redirect('usuarios:perfil')
    else:
        form = PerfilUsuarioForm(instance=request.user)
    return render(request, 'usuarios/perfil.html', {'form': form})

@login_required
def cambiar_contrasena(request):
    if request.method == 'POST':
        form = CambiarContrasenaForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Contraseña cambiada exitosamente.')
            return redirect('usuarios:perfil')
    else:
        form = CambiarContrasenaForm(request.user)
    return render(request, 'usuarios/cambiar_contrasena.html', {'form': form})

@login_required
@requiere_permiso('exportar_usuarios')
def exportar_usuarios_excel(request):
    # Aplicar los mismos filtros que en la lista
    search = request.GET.get('search', '')
    order_by = request.GET.get('order_by', 'username')
    
    usuarios = Usuario.objects.all()
    
    if search:
        usuarios = usuarios.filter(
            Q(username__icontains=search) | 
            Q(email__icontains=search) | 
            Q(first_name__icontains=search) | 
            Q(last_name__icontains=search) |
            Q(rol__icontains=search)
        )
    
    if order_by in ['username', '-username', 'email', '-email', 'date_joined', '-date_joined', 'rol', '-rol']:
        usuarios = usuarios.order_by(order_by)
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    
    # Encabezados
    headers = ['Usuario', 'Nombres', 'Apellidos', 'Email', 'Rol', 'Área', 'Estado', 'Fecha Registro']
    ws.append(headers)
    
    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
    
    # Datos
    for usuario in usuarios:
        ws.append([
            usuario.username,
            usuario.first_name,
            usuario.last_name,
            usuario.email,
            usuario.get_rol_display(),
            usuario.area or '',
            usuario.estado,
            usuario.date_joined.strftime('%d/%m/%Y')
        ])
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'usuarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response

@login_required
@requiere_permiso('editar', 'usuarios')
def gestionar_roles(request, pk):
    usuario = get_object_or_404(Usuario, pk=pk)
    if request.method == 'POST':
        # Actualizar rol si se proporciona
        nuevo_rol = request.POST.get('rol')
        if nuevo_rol and nuevo_rol in dict(Usuario.ROLES):
            usuario.rol = nuevo_rol
            # Solo aplicar permisos predeterminados si se cambió el rol
            if 'aplicar_permisos_rol' in request.POST:
                usuario.aplicar_permisos_por_rol()
        
        # Actualizar permisos granulares (excepto para administradores)
        if usuario.rol != 'ADMINISTRADOR':
            permisos_granulares = [
                'puede_ver_usuarios', 'puede_crear_usuarios', 'puede_editar_usuarios', 'puede_desactivar_usuarios',
                'puede_ver_productos', 'puede_crear_productos', 'puede_editar_productos', 'puede_desactivar_productos',
                'puede_ver_proveedores', 'puede_crear_proveedores', 'puede_editar_proveedores', 'puede_desactivar_proveedores',
                'puede_ver_inventario', 'puede_crear_inventario', 'puede_editar_inventario',
                'puede_exportar_usuarios', 'puede_exportar_productos', 'puede_exportar_proveedores', 'puede_exportar_inventario',
                'puede_ver_reportes'
            ]
            
            # Actualizar cada permiso basado en si está marcado en el formulario
            for permiso in permisos_granulares:
                valor = request.POST.get(permiso) == 'on'
                setattr(usuario, permiso, valor)
        
        usuario.save()
        messages.success(request, f'Permisos de "{usuario.username}" actualizados exitosamente.')
        return redirect('usuarios:usuarios_lista')
    
    # Preparar contexto con permisos organizados
    permisos_por_modulo = {
        'usuarios': [
            ('puede_ver_usuarios', 'Ver usuarios'),
            ('puede_crear_usuarios', 'Crear usuarios'),
            ('puede_editar_usuarios', 'Editar usuarios'),
            ('puede_desactivar_usuarios', 'Desactivar usuarios'),
            ('puede_exportar_usuarios', 'Exportar usuarios'),
        ],
        'productos': [
            ('puede_ver_productos', 'Ver productos'),
            ('puede_crear_productos', 'Crear productos'),
            ('puede_editar_productos', 'Editar productos'),
            ('puede_desactivar_productos', 'Desactivar productos'),
            ('puede_exportar_productos', 'Exportar productos'),
        ],
        'proveedores': [
            ('puede_ver_proveedores', 'Ver proveedores'),
            ('puede_crear_proveedores', 'Crear proveedores'),
            ('puede_editar_proveedores', 'Editar proveedores'),
            ('puede_desactivar_proveedores', 'Desactivar proveedores'),
            ('puede_exportar_proveedores', 'Exportar proveedores'),
        ],
        'inventario': [
            ('puede_ver_inventario', 'Ver inventario'),
            ('puede_crear_inventario', 'Crear movimientos'),
            ('puede_editar_inventario', 'Editar inventario'),
            ('puede_exportar_inventario', 'Exportar inventario'),
        ],
        'reportes': [
            ('puede_ver_reportes', 'Ver reportes'),
        ],
    }
    
    context = {
        'usuario': usuario,
        'roles_disponibles': Usuario.ROLES,
        'permisos_por_modulo': permisos_por_modulo,
    }
    return render(request, 'usuarios/gestionar_roles.html', context)

