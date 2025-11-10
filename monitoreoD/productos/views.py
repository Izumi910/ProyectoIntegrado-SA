from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from .models import Producto
from .forms import ProductoForm
from django.http import JsonResponse, HttpResponse
from usuarios.decorators import requiere_permiso
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

@login_required
def login_view(request):
    if request.method == "POST":
        usuario = request.POST.get("usuario")
        contraseña = request.POST.get("contraseña")
        user = authenticate(request, username=usuario, password=contraseña)
        if user is not None:
            login(request, user)
            return redirect("productos:lista")
        else:
            messages.error(request, "Usuario o contraseña incorrectos")
    return render(request, "productos/login.html")

@login_required
@requiere_permiso('ver', 'productos')
def lista_productos(request):
    # Búsqueda
    search = request.GET.get('search', '')
    # Ordenamiento
    order_by = request.GET.get('order_by', 'nombre')
    # Paginación
    per_page = request.GET.get('per_page', '10')
    
    productos = Producto.objects.all()
    
    # Aplicar búsqueda
    if search:
        productos = productos.filter(
            Q(nombre__icontains=search) | 
            Q(sku__icontains=search) | 
            Q(ean_upc__icontains=search) |
            Q(marca__icontains=search) |
            Q(categoria__nombre__icontains=search)
        )
    
    # Aplicar ordenamiento
    if order_by in ['nombre', '-nombre', 'sku', '-sku', 'precio_venta', '-precio_venta', 'stock', '-stock']:
        productos = productos.order_by(order_by)
    
    # Paginación
    paginator = Paginator(productos, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # AJAX response for live search
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        productos_data = []
        for producto in page_obj:
            productos_data.append({
                'id': producto.pk,
                'sku': producto.sku,
                'nombre': producto.nombre,
                'categoria': producto.categoria.nombre if producto.categoria else '',
                'marca': producto.marca or '-',
                'precio_venta': f"${int(producto.precio_venta) if producto.precio_venta and producto.precio_venta == int(producto.precio_venta) else producto.precio_venta or 0}",
                'stock': int(producto.stock) if producto.stock == int(producto.stock) else producto.stock,
                'stock_minimo': producto.stock_minimo,
                'estado': producto.estado,
                'estado_class': 'text-bg-success' if producto.estado == 'ACTIVO' else 'text-bg-secondary'
            })
        return JsonResponse({
            'productos': productos_data,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'page_number': page_obj.number,
            'num_pages': page_obj.paginator.num_pages
        })
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'order_by': order_by,
        'per_page': per_page,
    }
    
    return render(request, "productos/lista.html", context)

@login_required
def detalle_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    return render(request, "productos/detalle.html", {"producto": producto})

@login_required
@requiere_permiso('crear', 'productos')
def crear_producto(request):
    if request.method == "POST":
        form = ProductoForm(request.POST)
        if form.is_valid():
            producto = form.save(commit=False)
            # Asegurar que todos los campos se guarden
            producto.save()
            messages.success(request, f'Producto "{producto.nombre}" creado exitosamente.')
            return redirect("productos:productos_lista")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{form.fields[field].label}: {error}')
    else:
        form = ProductoForm()
    return render(request, "productos/crear.html", {"form": form})

@login_required
@requiere_permiso('editar', 'productos')
def editar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == "POST":
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            producto = form.save()
            messages.success(request, f'Producto "{producto.nombre}" actualizado exitosamente.')
            return redirect("productos:productos_lista")
    else:
        form = ProductoForm(instance=producto)
    return render(request, "productos/editar.html", {"form": form})

@login_required
@requiere_permiso('desactivar', 'productos')
def toggle_producto_estado(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == "POST":
        nombre_producto = producto.nombre
        nuevo_estado = 'INACTIVO' if producto.estado == 'ACTIVO' else 'ACTIVO'
        producto.estado = nuevo_estado
        producto.save()
        accion = 'desactivado' if nuevo_estado == 'INACTIVO' else 'activado'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'Producto "{nombre_producto}" {accion} exitosamente.'})
        messages.success(request, f'Producto "{nombre_producto}" {accion} exitosamente.')
        return redirect("productos:productos_lista")
    return render(request, "productos/toggle_estado.html", {"producto": producto})

@login_required
def panel_productos(request):
    productos = Producto.objects.all()
    return render(request, 'productos/panel.html', {'productos': productos})

@login_required
def inicio(request):
    visitas = request.session.get('visitas', 0)
    request.session['visitas'] = visitas + 1

    return render(request, 'productos/inicio.html', {'visitas': visitas})

@login_required
@requiere_permiso('exportar_productos')
def exportar_productos_excel(request):
    search = request.GET.get('search', '')
    order_by = request.GET.get('order_by', 'nombre')
    
    productos = Producto.objects.all()
    
    if search:
        productos = productos.filter(
            Q(nombre__icontains=search) | 
            Q(sku__icontains=search) | 
            Q(ean_upc__icontains=search) |
            Q(marca__icontains=search) |
            Q(categoria__nombre__icontains=search)
        )
    
    if order_by in ['nombre', '-nombre', 'sku', '-sku', 'precio_venta', '-precio_venta', 'stock', '-stock']:
        productos = productos.order_by(order_by)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    headers = ['SKU', 'EAN/UPC', 'Nombre', 'Categoría', 'Marca', 'Modelo', 'Descripción', 'Costo Estándar', 'Costo Promedio', 'Precio Venta', 'IVA %', 'Stock', 'Stock Mínimo', 'Stock Máximo', 'Punto Reorden', 'Lote', 'Fecha Vencimiento', 'Estado']
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
    
    for producto in productos:
        ws.append([
            producto.sku,
            producto.ean_upc or '',
            producto.nombre,
            producto.categoria.nombre if producto.categoria else '',
            producto.marca or '',
            producto.modelo or '',
            producto.descripcion or '',
            float(producto.costo_estandar) if producto.costo_estandar else 0,
            float(producto.costo_promedio) if producto.costo_promedio else 0,
            float(producto.precio_venta) if producto.precio_venta else 0,
            float(producto.impuesto_iva) if producto.impuesto_iva else 0,
            producto.stock,
            producto.stock_minimo,
            producto.stock_maximo or '',
            producto.punto_reorden or '',
            producto.lote or '',
            producto.fecha_vencimiento.strftime('%d/%m/%Y') if producto.fecha_vencimiento else '',
            producto.estado
        ])
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'productos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response







