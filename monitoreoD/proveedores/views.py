from django.shortcuts import render, get_object_or_404, redirect
from .models import Proveedor, ProveedorProducto
from .forms import ProveedorForm, ProveedorProductoForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from usuarios.decorators import requiere_permiso
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

@login_required
@requiere_permiso('ver', 'proveedores')
def lista_proveedores(request):
    search = request.GET.get('search', '')
    order_by = request.GET.get('order_by', 'razon_social')
    per_page = request.GET.get('per_page', '10')
    
    proveedores = Proveedor.objects.all()
    
    if search:
        proveedores = proveedores.filter(
            Q(razon_social__icontains=search) | 
            Q(rut_nif__icontains=search) | 
            Q(email__icontains=search) | 
            Q(ciudad__icontains=search)
        )
    
    if order_by in ['razon_social', '-razon_social', 'rut_nif', '-rut_nif', 'email', '-email', 'ciudad', '-ciudad']:
        proveedores = proveedores.order_by(order_by)
    
    paginator = Paginator(proveedores, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # AJAX response for live search
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        proveedores_data = []
        for proveedor in page_obj:
            proveedores_data.append({
                'id': proveedor.pk,
                'rut_nif': proveedor.rut_nif,
                'razon_social': proveedor.razon_social,
                'email': proveedor.email,
                'telefono': proveedor.telefono or '',
                'ciudad': proveedor.ciudad or '',
                'estado': proveedor.estado,
                'estado_class': 'text-bg-success' if proveedor.estado == 'ACTIVO' else 'text-bg-secondary'
            })
        return JsonResponse({
            'proveedores': proveedores_data,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'page_number': page_obj.number,
            'num_pages': page_obj.paginator.num_pages
        })
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'order_by': order_by,
        'per_page': per_page,
    }
    
    return render(request, 'proveedores/lista.html', context)

@login_required
@requiere_permiso('crear', 'proveedores')
def crear_proveedor(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            proveedor = form.save()
            messages.success(request, f'Proveedor "{proveedor.razon_social}" creado exitosamente.')
            return redirect('proveedores:proveedores_lista')
    else:
        form = ProveedorForm()
    return render(request, 'proveedores/crear.html', {'form': form})

@login_required
@requiere_permiso('editar', 'proveedores')
def editar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            proveedor = form.save()
            messages.success(request, f'Proveedor "{proveedor.razon_social}" actualizado exitosamente.')
            return redirect('proveedores:proveedores_lista')
    else:
        form = ProveedorForm(instance=proveedor)
    return render(request, 'proveedores/editar.html', {'form': form, 'proveedor': proveedor})

@login_required
@requiere_permiso('desactivar', 'proveedores')
def toggle_proveedor_estado(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        razon_social = proveedor.razon_social
        nuevo_estado = 'INACTIVO' if proveedor.estado == 'ACTIVO' else 'ACTIVO'
        proveedor.estado = nuevo_estado
        proveedor.save()
        accion = 'desactivado' if nuevo_estado == 'INACTIVO' else 'activado'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'Proveedor "{razon_social}" {accion} exitosamente.'})
        messages.success(request, f'Proveedor "{razon_social}" {accion} exitosamente.')
        return redirect('proveedores:proveedores_lista')
    return render(request, 'proveedores/toggle_estado.html', {'proveedor': proveedor})

@login_required
def detalle_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    return render(request, 'proveedores/detalle.html', {'proveedor': proveedor})

@login_required
@requiere_permiso('exportar_proveedores')
def exportar_proveedores_excel(request):
    search = request.GET.get('search', '')
    order_by = request.GET.get('order_by', 'razon_social')
    
    proveedores = Proveedor.objects.all()
    
    if search:
        proveedores = proveedores.filter(
            Q(razon_social__icontains=search) | 
            Q(rut_nif__icontains=search) | 
            Q(email__icontains=search) | 
            Q(ciudad__icontains=search)
        )
    
    if order_by in ['razon_social', '-razon_social', 'rut_nif', '-rut_nif', 'email', '-email', 'ciudad', '-ciudad']:
        proveedores = proveedores.order_by(order_by)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    
    headers = ['RUT', 'Razón Social', 'Email', 'Teléfono', 'Ciudad', 'Estado']
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
    
    for proveedor in proveedores:
        ws.append([
            proveedor.rut_nif,
            proveedor.razon_social,
            proveedor.email,
            proveedor.telefono or '',
            proveedor.ciudad or '',
            proveedor.estado
        ])
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'proveedores_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


@login_required
def lista_proveedor_productos(request):
    relaciones = ProveedorProducto.objects.all()
    return render(request, 'proveedores/lista_proveedor_productos.html', {'relaciones': relaciones})

@login_required
def crear_proveedor_producto(request):
    if request.method == 'POST':
        form = ProveedorProductoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('proveedores:lista_proveedor_productos')
    else:
        form = ProveedorProductoForm()
    return render(request, 'proveedores/crear_proveedor_producto.html', {'form': form})

@login_required
def editar_proveedor_producto(request, pk):
    relacion = get_object_or_404(ProveedorProducto, pk=pk)
    if request.method == 'POST':
        form = ProveedorProductoForm(request.POST, instance=relacion)
        if form.is_valid():
            form.save()
            return redirect('proveedores:lista_proveedor_productos')